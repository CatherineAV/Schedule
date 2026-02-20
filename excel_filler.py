from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import datetime


class ExcelFiller:
    """
    Заполняет шаблон Excel данными из сгенерированного расписания
    """

    def __init__(self, db_ops=None):
        self.db_ops = db_ops

        # Соответствие дней недели (коды из schedule_data -> индексы)
        self.day_mapping = {
            'пн': 0,
            'вт': 1,
            'ср': 2,
            'чт': 3,
            'пт': 4,
            'сб': 5
        }

        # Начальная строка для данных (после заголовков)
        self.start_row = 14

        # Количество уроков в день
        self.lessons_per_day = 11

        # Кэш цветов территорий
        self.territory_colors = {}

    def load_territory_colors(self):
        """
        Загружает цвета территорий из базы данных
        """
        if not self.db_ops:
            return

        territories = self.db_ops.get_territories()
        for territory in territories:
            name = territory['Территория']
            color = territory.get('Цвет', '#FFFFFF')
            # Убираем # если есть
            if color.startswith('#'):
                color = color[1:]
            self.territory_colors[name] = color

    def fill_schedule(self, template_path: str, schedule_data: Dict,
                      group_structure: List[Dict], group_names: Dict[int, Dict],
                      output_path: str = None) -> str:
        """
        Заполняет шаблон данными из расписания
        group_names: {group_id: {'name': str, 'subgroup': str, 'full_name': str}}
        """
        # Загружаем цвета территорий
        self.load_territory_colors()

        # Загружаем шаблон
        wb = load_workbook(template_path)
        ws = wb.active

        # Создаем соответствие между группами и колонками в Excel
        group_columns = self._map_groups_to_columns(ws, group_structure, group_names)

        # Заполняем расписание
        self._fill_lessons(ws, schedule_data, group_columns, group_names)

        # Сохраняем результат
        if output_path:
            wb.save(output_path)
            return output_path
        else:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"filled_schedule_{timestamp}.xlsx"
            wb.save(output_path)
            return output_path

    def _map_groups_to_columns(self, ws, group_structure: List[Dict],
                               group_names: Dict[int, Dict]) -> Dict[int, int]:
        """
        Сопоставляет ID групп с колонками в Excel
        Возвращает словарь: {group_id: column_index}
        """
        group_columns = {}

        print("=" * 50)
        print("Поиск соответствия групп и колонок:")

        # Сначала построим структуру групп из Excel
        excel_groups = {}  # {group_name: {start_col: int, end_col: int, subgroups: dict}}

        col = 4
        while col < 100:
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}12']

            if not cell.value:
                # Если нет названия группы, проверяем следующую колонку
                col += 1
                continue

            group_name = str(cell.value).strip()

            # Проверяем, есть ли подгруппы в строке 13
            has_subgroups = False
            subgroups = {}

            # Смотрим, сколько колонок занимает эта группа
            start_col = col
            end_col = start_col

            # Проверяем следующую колонку
            next_col = col + 1
            next_col_letter = get_column_letter(next_col)
            next_cell = ws[f'{next_col_letter}12']

            # Если следующая колонка пустая или это новая группа
            if not next_cell.value or next_cell.value != group_name:
                # Группа занимает одну колонку
                sub_cell = ws[f'{col_letter}13']
                if sub_cell.value and sub_cell.value != '—':
                    has_subgroups = True
                    subgroups[start_col] = str(sub_cell.value).strip()
                end_col = start_col
                col = next_col
            else:
                # Группа занимает несколько колонок (с подгруппами)
                while next_col < 100:
                    next_col_letter = get_column_letter(next_col)
                    next_cell = ws[f'{next_col_letter}12']

                    if next_cell.value and next_cell.value == group_name:
                        # Это продолжение той же группы
                        sub_cell = ws[f'{next_col_letter}13']
                        if sub_cell.value:
                            subgroups[next_col] = str(sub_cell.value).strip()
                        end_col = next_col
                        next_col += 1
                    else:
                        break

                has_subgroups = True
                col = next_col

            excel_groups[group_name] = {
                'start_col': start_col,
                'end_col': end_col,
                'subgroups': subgroups,
                'has_subgroups': has_subgroups
            }

            print(f"Excel группа: '{group_name}' колонки {start_col}-{end_col}, подгруппы: {subgroups}")

        print("\nГруппы из базы данных:")
        for group_id, info in group_names.items():
            print(f"  ID {group_id}: '{info['full_name']}' (имя: '{info['name']}', подгруппа: '{info['subgroup']}')")

        # Сопоставляем группы из БД с колонками в Excel
        for group_id, info in group_names.items():
            group_name = info['name']
            subgroup = info['subgroup']

            found = False

            # Ищем группу в Excel
            if group_name in excel_groups:
                excel_group = excel_groups[group_name]

                if subgroup == 'Нет':
                    # Группа без подгруппы - должна быть одна колонка
                    if not excel_group['has_subgroups']:
                        group_columns[group_id] = excel_group['start_col']
                        print(
                            f"✓ Группа {group_id} ('{group_name}') без подгруппы в колонке {excel_group['start_col']}")
                        found = True
                else:
                    # Группа с подгруппой - ищем подходящую подгруппу
                    for sub_col, sub_value in excel_group['subgroups'].items():
                        # Проверяем соответствие подгруппы
                        if subgroup in ['1', '2', '3'] and f'{subgroup} п/гр' in sub_value:
                            group_columns[group_id] = sub_col
                            print(
                                f"✓ Группа {group_id} ('{group_name}') с подгруппой '{subgroup}' в колонке {sub_col} (подпись '{sub_value}')")
                            found = True
                            break
                        elif subgroup == 'Женская' and ('Женская' in sub_value or 'Жен' in sub_value):
                            group_columns[group_id] = sub_col
                            print(
                                f"✓ Группа {group_id} ('{group_name}') с подгруппой '{subgroup}' в колонке {sub_col} (подпись '{sub_value}')")
                            found = True
                            break
                        elif subgroup == 'Мужская' and ('Мужская' in sub_value or 'Муж' in sub_value):
                            group_columns[group_id] = sub_col
                            print(
                                f"✓ Группа {group_id} ('{group_name}') с подгруппой '{subgroup}' в колонке {sub_col} (подпись '{sub_value}')")
                            found = True
                            break
                        elif subgroup == 'Кукольники' and ('Кукольники' in sub_value or 'Кук' in sub_value):
                            group_columns[group_id] = sub_col
                            print(
                                f"✓ Группа {group_id} ('{group_name}') с подгруппой '{subgroup}' в колонке {sub_col} (подпись '{sub_value}')")
                            found = True
                            break
                        elif subgroup == 'Бутафоры' and ('Бутафоры' in sub_value or 'Бут' in sub_value):
                            group_columns[group_id] = sub_col
                            print(
                                f"✓ Группа {group_id} ('{group_name}') с подгруппой '{subgroup}' в колонке {sub_col} (подпись '{sub_value}')")
                            found = True
                            break

            if not found:
                print(f"✗ Не найдена колонка для группы {group_id} ('{info['full_name']}')")

        print("=" * 50)
        return group_columns

    def _find_group_column(self, group_name: str, subgroup: str,
                           group_columns: Dict[str, int]) -> Optional[int]:
        """
        Находит колонку для конкретной группы и подгруппы
        """
        # Формируем полное имя группы
        if subgroup and subgroup != 'Нет':
            full_name = f"{group_name} ({subgroup})"
        else:
            full_name = group_name

        # Ищем точное совпадение
        for name, col in group_columns.items():
            if name == full_name:
                return col

        # Если не нашли, ищем частичное совпадение
        for name, col in group_columns.items():
            if group_name in name:
                if subgroup == 'Нет' and '(' not in name:
                    return col
                elif subgroup != 'Нет' and subgroup in name:
                    return col

        return None

    def _fill_lessons(self, ws, schedule_data: Dict, group_columns: Dict[int, int],
                      group_names: Dict[int, Dict]):
        """
        Заполняет ячейки расписания
        """
        print("\nЗаполнение расписания:")

        # Получаем информацию об объединенных ячейках
        merged_cells = ws.merged_cells.ranges

        for group_id, group_schedule in schedule_data.items():
            # Проверяем, есть ли колонка для этой группы
            if group_id not in group_columns:
                print(f"  Пропускаем группу {group_id} - нет колонки")
                continue

            col = group_columns[group_id]
            col_letter = get_column_letter(col)

            group_info = group_names.get(group_id, {})
            group_full_name = group_info.get('full_name', f"Группа {group_id}")

            print(f"\n  Группа {group_id} ('{group_full_name}') в колонке {col}")

            lesson_count = 0

            # Заполняем каждый день
            for day_code, day_lessons in group_schedule.items():
                if day_code not in self.day_mapping:
                    continue

                day_idx = self.day_mapping[day_code]
                day_name = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб'][day_idx]

                for lesson_idx, lesson_data in enumerate(day_lessons):
                    if lesson_data is None:
                        continue

                    lesson_count += 1

                    # Вычисляем строку в Excel
                    row = self.start_row + (day_idx * self.lessons_per_day) + lesson_idx

                    # Проверяем, не является ли эта ячейка частью объединенного диапазона
                    cell_coord = f'{col_letter}{row}'
                    is_merged = False
                    merged_cell_value = None

                    for merged_range in merged_cells:
                        if cell_coord in merged_range:
                            is_merged = True
                            # Получаем левую верхнюю ячейку объединенного диапазона
                            top_left = str(merged_range).split(':')[0]
                            merged_cell_value = ws[top_left].value
                            break

                    # Если ячейка объединена и содержит "День самообразования", пропускаем
                    if is_merged and merged_cell_value and "День самообразования" in str(merged_cell_value):
                        print(
                            f"    Урок {lesson_idx + 1} в {day_name} - пропущен (день самообразования в объединенной ячейке)")
                        continue

                    # Если ячейка объединена, но не содержит день самообразования, это проблема
                    if is_merged:
                        print(f"    ВНИМАНИЕ: ячейка {cell_coord} объединена, но не содержит день самообразования")
                        continue

                    # Формируем текст для ячейки
                    cell_text = self._format_lesson_text(lesson_data)

                    # Записываем в ячейку
                    cell = ws[f'{col_letter}{row}']
                    cell.value = cell_text

                    # Применяем стиль с учетом территории
                    territory = lesson_data.get('territory')
                    self._apply_lesson_style(cell, territory)

                    subject = lesson_data.get('subject', '')
                    print(f"    Урок {lesson_idx + 1} в {day_name}: '{subject}' -> ячейка {col_letter}{row}")

            print(f"  Всего уроков для группы {group_id}: {lesson_count}")

    def _format_lesson_text(self, lesson_data: Dict) -> str:
        """
        Форматирует текст для ячейки расписания с учетом ограничений по ширине
        """
        subject = lesson_data.get('subject', '')
        teacher = lesson_data.get('teacher', '')
        classroom = lesson_data.get('classroom', '')

        # Сокращаем предмет, если он слишком длинный
        if len(subject) > 15:
            subject = subject[:12] + '...'

        parts = [subject]

        if teacher:
            # Сокращаем ФИО до инициалов
            teacher_short = self._shorten_teacher_name(teacher)
            # Еще больше сокращаем, если нужно
            if len(teacher_short) > 10:
                teacher_short = teacher_short[:8] + '...'
            parts.append(teacher_short)

        if classroom:
            # Сокращаем номер кабинета
            if len(classroom) > 8:
                classroom = classroom[:6] + '...'
            parts.append(f"каб.{classroom}")

        # Добавляем отметку о четности недели (сокращенно)
        week_parity = lesson_data.get('week_parity')
        if week_parity and week_parity != 'обе':
            # Сокращаем до "в" или "н"
            short_parity = "в" if week_parity == 'верхняя' else "н"
            parts.append(f"({short_parity})")

        return '\n'.join(parts)

    def _shorten_teacher_name(self, full_name: str) -> str:
        """
        Сокращает ФИО до формата "Фамилия И.О."
        """
        if not full_name:
            return ""

        parts = full_name.split()
        if len(parts) >= 3:
            return f"{parts[0]} {parts[1][0]}.{parts[2][0]}."
        elif len(parts) == 2:
            return f"{parts[0]} {parts[1][0]}."
        else:
            return parts[0]

    def _apply_lesson_style(self, cell, territory: str = None):
        """
        Применяет стиль к ячейке без изменения размеров
        """
        # Устанавливаем выравнивание
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=False,  # Отключаем перенос текста
            shrink_to_fit=False  # Не сжимаем текст
        )

        # Устанавливаем шрифт
        cell.font = Font(size=10)

        # Окрашиваем в цвет территории, если она есть
        if territory and territory in self.territory_colors:
            color = self.territory_colors[territory]
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )

    def _apply_styles(self, ws):
        """
        Применяет финальные стили ко всей таблице
        """
        # Обновляем границы для всей таблицы
        max_row = ws.max_row
        max_col = ws.max_column

        for row in range(12, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value or (row >= 14 and col >= 4):
                    if not cell.border:
                        cell.border = self.thin_border
