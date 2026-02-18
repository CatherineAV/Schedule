import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from database.operations import DBOperations
from database.settings_manager import SettingsManager


class SimpleTemplateGenerator:
    def __init__(self, db_ops: DBOperations):
        self.db_ops = db_ops
        self.settings_manager = SettingsManager(db_ops)

        self.lesson_times = [
            "9.00 - 9.45", "9.55 - 10.40", "10.50 - 11.35",
            "11.45 - 12.30", "13.00-13.45", "14.15-15.00",
            "15.15 - 16.00", "16.10 - 16.55", "17.05 - 17.50",
            "18.00 - 18.45", "18.55 - 19.40"
        ]

        self.days_vertical = {
            'пн': ['п', 'о', 'н', 'е', 'д', 'е', 'л', 'ь', 'н', 'и', 'к'],
            'вт': ['в', 'т', 'о', 'р', 'н', 'и', 'к'],
            'ср': ['с', 'р', 'е', 'д', 'а'],
            'чт': ['ч', 'е', 'т', 'в', 'е', 'р', 'г'],
            'пт': ['п', 'я', 'т', 'н', 'и', 'ц', 'а'],
            'сб': ['с', 'у', 'б', 'б', 'о', 'т', 'а']
        }

        self.days_order = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб']

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_template_with_groups(self, output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание I семестр"

        active_groups = self._get_active_groups()
        group_structure = self._build_group_structure(active_groups)

        # ========== ШАПКА ДОКУМЕНТА ==========
        self._create_document_header(ws)

        # ========== ЗАГОЛОВОК ТАБЛИЦЫ ==========
        self._create_table_header(ws, group_structure)

        # ========== ДНИ И УРОКИ с учетом самообразования ==========
        self._create_days_and_lessons(ws, group_structure)

        # ========== ПОДВАЛ ==========
        # self._create_footer(ws)

        # ========== НАСТРОЙКА СТИЛЕЙ ==========
        self._apply_styles(ws)
        self._adjust_column_widths(ws)

        # Сохранение
        wb.save(output_path)
        return os.path.abspath(output_path)

    def _get_active_groups(self) -> List[Dict[str, Any]]:

        groups = self.settings_manager.get_groups_with_exclusion_and_order()
        active_groups = [g for g in groups if not g['Исключена']]
        active_groups.sort(key=lambda x: x['Порядок'])

        return active_groups

    def _build_group_structure(self, groups: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        grouped = {}

        for group in groups:
            name = group['Группа']
            subgroup = group.get('Подгруппа', 'Нет')

            if name not in grouped:
                grouped[name] = {
                    'name': name,
                    'subgroups': [],
                    'has_subgroups': False,
                    'self_education': group.get('Самообразование'),
                    'important_talks': group.get('Разговоры о важном', 0)
                }

            if subgroup != 'Нет' and subgroup != 'None':
                grouped[name]['subgroups'].append(subgroup)
                grouped[name]['has_subgroups'] = True

        structure = []
        for name, group_info in grouped.items():
            if group_info['has_subgroups']:
                subgroups = group_info['subgroups']

                if 'ХКО' in name.upper():
                    order = {'Женская': 1, 'Мужская': 2}
                    subgroups.sort(key=lambda x: order.get(x, 999))
                elif 'ХБО' in name.upper():
                    order = {'Кукольники': 1, 'Бутафоры': 2}
                    subgroups.sort(key=lambda x: order.get(x, 999))
                else:
                    subgroups.sort(key=lambda x: int(x) if x.isdigit() else 999)

                group_info['subgroups'] = subgroups
                structure.append(group_info)
            else:
                structure.append({
                    'name': name,
                    'subgroups': [],
                    'has_subgroups': False,
                    'self_education': group_info['self_education']
                })

        return structure

    def _create_document_header(self, ws):

        ws.merge_cells('A1:CN1')
        ws['A1'] = 'Расписание занятий в I полугодии 2025/26 учебного года'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('AP3:AP5')
        ws['AP3'] = 'УТВЕРЖДАЮ'
        ws['AP3'].font = Font(bold=True)
        ws['AP3'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('AP6:AP7')
        ws['AP6'] = 'Директор ГБПОУ г. Москвы "ТХТК"'
        ws['AP6'].font = Font(bold=True)
        ws['AP6'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('AP8:AP9')
        ws['AP8'] = '____________Подбуртная Н.Н.'
        ws['AP8'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('AP10:AP11')
        ws['AP10'] = '"______"  _____________  20____ г.'
        ws['AP10'].alignment = Alignment(horizontal='center', vertical='center')

    def _create_table_header(self, ws, group_structure: List[Dict[str, Any]]):

        header_row = 12
        subgroup_row = 13

        fixed_columns = ['A', 'B', 'C']

        for col in fixed_columns:
            merge_range = f'{col}{header_row}:{col}{subgroup_row}'
            ws.merge_cells(merge_range)

            if col == 'A':
                ws[f'{col}{header_row}'] = 'Дни'
            elif col == 'B':
                ws[f'{col}{header_row}'] = 'Уроки'
            elif col == 'C':
                ws[f'{col}{header_row}'] = 'Расписание звонков'

            cell = ws[f'{col}{header_row}']
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_col = 4

        for group in group_structure:
            if group['has_subgroups']:
                subgroup_count = len(group['subgroups'])

                start_col = current_col
                end_col = current_col + subgroup_count - 1

                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)

                merge_range = f'{start_letter}{header_row}:{end_letter}{header_row}'
                ws.merge_cells(merge_range)

                group_cell = ws[f'{start_letter}{header_row}']
                group_cell.value = group['name']
                group_cell.font = Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for i, subgroup in enumerate(group['subgroups']):
                    col = current_col + i
                    col_letter = get_column_letter(col)

                    sub_cell = ws[f'{col_letter}{subgroup_row}']

                    if subgroup in ['1', '2', '3']:
                        sub_cell.value = f'{subgroup} п/гр'
                    elif subgroup == 'Женская':
                        sub_cell.value = 'Женская'
                    elif subgroup == 'Мужская':
                        sub_cell.value = 'Мужская'
                    elif subgroup == 'Кукольники':
                        sub_cell.value = 'Кукольники'
                    elif subgroup == 'Бутафоры':
                        sub_cell.value = 'Бутафоры'
                    else:
                        sub_cell.value = subgroup

                    sub_cell.font = Font(size=9)
                    sub_cell.alignment = Alignment(horizontal='center', vertical='center')

                current_col += subgroup_count

            else:
                col_letter = get_column_letter(current_col)

                merge_range = f'{col_letter}{header_row}:{col_letter}{subgroup_row}'
                ws.merge_cells(merge_range)

                group_cell = ws[f'{col_letter}{header_row}']
                group_cell.value = group['name']
                group_cell.font = Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_col += 1

    def _create_days_and_lessons(self, ws, group_structure: List[Dict[str, Any]]):

        start_row = 14

        day_mapping = {
            'Пн': 'пн', 'ПН': 'пн', 'пн': 'пн',
            'Вт': 'вт', 'ВТ': 'вт', 'вт': 'вт',
            'Ср': 'ср', 'СР': 'ср', 'ср': 'ср',
            'Чт': 'чт', 'ЧТ': 'чт', 'чт': 'чт',
            'Пт': 'пт', 'ПТ': 'пт', 'пт': 'пт',
            'Сб': 'сб', 'СБ': 'сб', 'сб': 'сб'
        }

        for day_idx, day_code in enumerate(self.days_order):
            day_start_row = start_row + (day_idx * 11)
            day_end_row = day_start_row + 10

            merge_range = f'A{day_start_row}:A{day_end_row}'
            ws.merge_cells(merge_range)

            day_letters = self.days_vertical[day_code]
            vertical_text = '\n'.join(day_letters).upper()

            day_cell = ws[f'A{day_start_row}']
            day_cell.value = vertical_text
            day_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            day_cell.font = Font(bold=True)

            current_col = 4

            for group in group_structure:
                self_education = group.get('self_education')

                if self_education and self_education != 'Нет' and self_education != 'None':
                    self_ed_code = day_mapping.get(self_education)

                    if self_ed_code == day_code:
                        if group['has_subgroups']:
                            subgroup_count = len(group['subgroups'])

                            start_col = current_col
                            end_col = current_col + subgroup_count - 1

                            start_letter = get_column_letter(start_col)
                            end_letter = get_column_letter(end_col)

                            merge_range = f'{start_letter}{day_start_row}:{end_letter}{day_end_row}'
                            ws.merge_cells(merge_range)

                            cell = ws[f'{start_letter}{day_start_row}']
                            cell.value = "День самообразования"
                            cell.alignment = Alignment(
                                horizontal='center',
                                vertical='center',
                                wrap_text=True
                            )
                            cell.font = Font(bold=True, size=10)

                            current_col += subgroup_count
                        else:
                            col_letter = get_column_letter(current_col)

                            merge_range = f'{col_letter}{day_start_row}:{col_letter}{day_end_row}'
                            ws.merge_cells(merge_range)

                            cell = ws[f'{col_letter}{day_start_row}']
                            cell.value = "День самообразования"
                            cell.alignment = Alignment(
                                horizontal='center',
                                vertical='center',
                                wrap_text=True
                            )
                            cell.font = Font(italic=True, bold=True, size=10)

                            current_col += 1
                    else:
                        if group['has_subgroups']:
                            current_col += len(group['subgroups'])
                        else:
                            current_col += 1
                else:
                    if group['has_subgroups']:
                        current_col += len(group['subgroups'])
                    else:
                        current_col += 1

            for lesson_idx in range(11):
                row = day_start_row + lesson_idx

                ws.cell(row=row, column=2, value=lesson_idx + 1)
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(row=row, column=3, value=self.lesson_times[lesson_idx])
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

    # def _create_footer(self, ws):
    #     last_row = ws.max_row + 3
    #
    #     ws.merge_cells(f'CH{last_row}:CN{last_row}')
    #     ws[f'CH{last_row}'] = 'Заместитель директора по УР________________Соломина И.Д.'
    #     ws[f'CH{last_row}'].alignment = Alignment(horizontal='right')

    def _apply_styles(self, ws):
        last_row = 14 + (6 * 11) - 1

        last_col = 3

        for col in range(4, 100):
            col_letter = get_column_letter(col)

            cell_row12 = ws[f'{col_letter}12']
            cell_row13 = ws[f'{col_letter}13']

            if not cell_row12.value and not cell_row13.value:
                last_col = col - 1
                break

        if last_col == 3:
            has_any_group = False
            for col in range(4, 30):
                col_letter = get_column_letter(col)
                if ws[f'{col_letter}12'].value or ws[f'{col_letter}13'].value:
                    has_any_group = True
                    last_col = col

            if not has_any_group:
                last_col = 10

        for row in range(12, last_row + 1):
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)

                if col == 1:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                else:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

        for day_idx in range(6):
            day_start_row = 14 + (day_idx * 11)
            day_end_row = day_start_row + 10

            for row in range(day_start_row, day_end_row + 1):
                cell = ws.cell(row=row, column=1)

                if row == day_start_row:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style=None)
                    )
                elif row == day_end_row:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style=None),
                        bottom=Side(style='thin')
                    )
                else:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style=None),
                        bottom=Side(style=None)
                    )

    def _adjust_column_widths(self, ws):

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 15

        for col in range(4, 50):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 12